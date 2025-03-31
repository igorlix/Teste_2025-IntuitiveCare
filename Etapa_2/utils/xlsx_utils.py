from openpyxl import load_workbook
import pandas as pd
from utils.data_processing import normalizar_texto

def ler_xlsx_corretamente(xlsx_path):
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active

    cabecalho = []
    for row in ws.iter_rows(min_row=4, max_row=5, min_col=1, max_col=13):
        cabecalho.append([cell.value for cell in row])

    nomes_colunas = []
    for col_idx in range(13):
        parte1 = str(cabecalho[0][col_idx] or "").strip()
        parte2 = str(cabecalho[1][col_idx] or "").strip()
        nomes_colunas.append(f"{parte1} {parte2}".strip())

    dados = []
    for row in ws.iter_rows(min_row=6, max_row=3397, min_col=1, max_col=13):
        dados.append([cell.value for cell in row])

    df = pd.DataFrame(dados, columns=nomes_colunas)
    df = df.dropna(how='all')

    for col in df.columns:
        df[col] = df[col].apply(normalizar_texto)

    return df.fillna('None')